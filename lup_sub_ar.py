#!/usr/local/bin/python3
'''
Esta es la rama principal del script
Imprime los resultados en un excel y maneja los dispo:

       Android Mobile : ADR M \n"
       iOS Mobile : IOS M \n"
       WEB : WEB \n"
       Claro TV TATA o IPTV/AOSP : Iptv \n"
       Roku : Rk \n"


'''
import requests
from requests.packages.urllib3.exceptions import InsecureRequestWarning
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
import os
import time
import sys

# Deshabilitar advertencias de solicitudes inseguras
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

class EpgCategory:
    def __init__(self, base_url, authpn, authpt, device_id, device_category, device_model, device_type, device_so, node_id):
        self.base_url = base_url
        self.authpn = authpn
        self.authpt = authpt
        self.device_id = device_id
        self.device_category = device_category
        self.device_model = device_model
        self.device_type = device_type
        self.device_so = device_so
        self.node_id = node_id

    def obtener_menu_id(self, subregion):
        url = f"{self.base_url}/version"
        params = {
            "authpn": self.authpn,
            "authpt": self.authpt,
            "device_id": self.device_id,
            "device_category": self.device_category,
            "device_model": self.device_model,
            "device_type": self.device_type,
            "device_so": self.device_so,
            "format": "json",
            "device_manufacturer": "generic",
            "api_version": "v5.93",
            "region": "argentina",
            "subregion": subregion
        }
        response = requests.get(url, params=params, verify=False)
        data = response.json()

        if "response" in data and isinstance(data["response"], dict):
            if "epg_version" in data["response"]:
                return data["response"]["epg_version"]
        return None

    def obtener_lineup(self, epg_version, subregion):
        url = f"{self.base_url}/lineup"
        params = {
            "authpn": self.authpn,
            "authpt": self.authpt,
            "device_id": self.device_id,
            "device_category": self.device_category,
            "device_model": self.device_model,
            "device_type": self.device_type,
            "device_so": self.device_so,
            "format": "json",
            "device_manufacturer": "generic",
            "api_version": "v5.93",
            "region": "argentina",
            "subregion": subregion,
            "epg_version": epg_version,
            "from": 0,
            "quantity": 2000,
            "metadata": "reduced"
        }
        response = requests.get(url, params=params, verify=False)
        data = response.json()

        if "response" in data and isinstance(data["response"], dict):
            if "channels" in data["response"] and "total" in data["response"]:
                return data["response"]["channels"], data["response"]["total"]
        return None, 0

def write_to_excel(channels, total, file_name='epg_data.xlsx'):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "EPG Data"

    # Definir estilos
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Escribir cabecera
    headers = ["Number", "Name", "Image"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    total_channels = len(channels)
    for index, channel in enumerate(channels, start=2):
        sheet.cell(row=index, column=1, value=channel['number']).border = border
        sheet.cell(row=index, column=2, value=channel['name']).border = border
        sheet.cell(row=index, column=3, value=channel['image']).border = border

        # Calcular y mostrar el porcentaje de progreso
        progress = (index - 1) / total_channels * 100
        sys.stdout.write(f"\rProgreso: {progress:.2f}% completado")
        sys.stdout.flush()
        time.sleep(0.01)  # Para simular tiempo de procesamiento

    sheet.cell(row=total_channels + 2, column=1, value="Total de canales:").font = Font(bold=True)
    sheet.cell(row=total_channels + 2, column=2, value=total).border = border

    # Guardar el archivo en la carpeta de descargas
    downloads_folder = os.path.expanduser("~/Downloads")
    file_path = os.path.join(downloads_folder, file_name)
    workbook.save(file_path)
    print(f"\nArchivo guardado en {file_path}")

if __name__ == "__main__":
    device_type_input = input("Android Mobile : ADR M \n"
                              "iOS Mobile : IOS M \n"
                              "WEB : WEB \n"
                              "Claro TV TATA o IPTV/AOSP : Iptv \n"
                              "Roku : Rk \n"
                              "Ingrese el dispositivo a consultar: ").strip().lower()
    subregion = input("Ingrese la subregion: ").strip()

    if device_type_input == "web":
        # Datos de autenticación y del dispositivo Web (OTT)
        base_url = "https://mfwkweb-api.clarovideo.net/services/epg"
        authpn = "webclient"
        authpt = "tfg1h3j4k6fd7"
        device_id = "web"
        device_category = "web"
        device_model = "web"
        device_type = "web"
        device_so = "Chrome"
        node_id = "19442"

    elif device_type_input == "iptv":
        base_url = "https://mfwkstbzte-api.clarovideo.net/services/epg"
        authpn = "tataelxsi"
        authpt = "vofee7ohhecai"
        device_id = "ZTEATV41200438593"
        device_category = "stb"
        device_model = "B866V2_V1_0_0"
        device_type = "ptv"
        device_so = "Android 12"
        node_id = "19083"

    elif device_type_input == "adr m":
        base_url = "https://mfwkmobileandroid-api.clarovideo.net/services/epg"
        authpn = "amco"
        authpt = "12e4i8l6a581a"
        device_id = "50e940ca-9ffa-4e91-a184-a4878e41238e"
        device_category = "mobile"
        device_model = "android"
        device_type = "SM-G970F"
        device_so = "Android 2013"
        node_id = "19442"

    elif device_type_input == "ios m":
        base_url = "https://mfwkmobileios-api.clarovideo.net/services/epg"
        authpn = "amco"
        authpt = "12e4i8l6a581a"
        device_id = "467C066A-F668-41A8-AC58-404B122EA991"
        device_category = "mobile"
        device_model = "aapl"
        device_type = "iPhone"
        device_so = "iOS 2015.0"
        node_id = "19442"

    elif device_type_input == "rk":
        base_url = "https://mfwkstbroku-api.clarovideo.net/services/epg"
        authpn = "roku"
        authpt = "IdbIIWeFzYdy"
        device_id = "f7785395-3dc0-5ca4-b2bd-b4e6346221e3"
        device_category = "stb"
        device_model = "generic"
        device_type = "generic"
        device_so = ""
        node_id = "19442"

    else:
        print("Tipo de dispositivo no reconocido. Verifique que sea ingresado correctamente.")
        exit()

    EPG_client = EpgCategory(base_url, authpn, authpt, device_id, device_category, device_model, device_type, device_so, node_id)

    menu_id = EPG_client.obtener_menu_id(subregion)

    if menu_id:
        canales, total = EPG_client.obtener_lineup(menu_id, subregion)
        if canales:
            # Construir el nombre del archivo
            file_name = f"{device_type_input}_epg_data_{subregion}.xlsx"
            write_to_excel(canales, total, file_name)
            print(f"\nDatos guardados en {file_name}.")
        else:
            print("No se encontraron canales.")
            for canal in canales:
                print(f"Channel: ")
                print(f"\tNumber: {canal['number']}")
                print(f"\tName: {canal['name']}")
                print(f"\tImage: {canal['image']}")
        print(f"\nTotal de canales: {total}")
    else:
        print(f"No se encontró la subregion '{subregion}' en la región de argentina.")

