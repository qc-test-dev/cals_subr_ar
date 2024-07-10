#!/usr/bin/python3
import requests
from openpyxl import load_workbook
from data_info import *
from login import login_cv
from epg_update import epg_version_update
class BrfCal:
    def __init__(self,row):
        for i, header in enumerate(inputs_headers):
            setattr(self,header,row[i])
            #print(header,row[i])
    
    def login(self):
        user=self.user
        password=self.password
        region=self.region
        dispositivo=self.device_id
        url_silo=self.url_silo
        device_id=self.device_id
        device_model = self.device_model
        device_type=self.device_type
        device_so=self.device_so
        device_category=self.device_category
        device_manufacturer=self.device_manufacturer
        try:
            datos_login=login_cv(user,password,region,dispositivo,url_silo,device_id,device_model,device_type,device_so,device_category,device_manufacturer)
            if datos_login:
                try:
                    self.hks= datos_login['session_stringvalue']
                    self.user_id=datos_login['user_id']
                    self.paymentMethods=datos_login['paymentMethods']
                    #self.user_hash=datos_login['user_hash']
                    print(self.hks,self.user_id) 
                except Exception  as e:
                     print("error en esta fase")
                return (self.hks,self.user_id,self.paymentMethods)
            else:
                self.user_id="credenciales invalidas, se validó con hks generico"
                self.hks='0'
                #self.user_hash='0' 
                #print(self.hks,self.user_id)    
                return (self.hks,self.user_id)
        except Exception as e:
                self.user_id="credenciales invalidas, se valdio con hks generico" 
                self.hks='0'
                #self.user_hash='0'
                print(self.hks,self.user_id,e) 
                return (self.hks,self.user_id)
        

    def epgupdate(self):
        user=self.user
        password=self.password
        region=self.region
        dispositivo=self.device_id
        url_silo=self.url_silo
        device_id=self.device_id
        device_model = self.device_model
        device_type=self.device_type
        device_so=self.device_so
        device_category=self.device_category
        device_manufacturer=self.device_manufacturer
        hks=self.hks
        try:
            datos_epg=epg_version_update(user,password,region,url_silo,device_id,device_model,device_type,device_so,device_category,device_manufacturer,hks)
            
            if datos_epg:
                try:
                    self.epg_version= datos_epg['epg_version']  
                    print(self.epg_version) 
                except Exception  as e:
                    print("error en esta fase")
            else:
                self.epg_version="epg invalido"  
                return (self.epg_version)
        except Exception as e:
                
                self.epg_version='0' 
                return (self.epg_version)  


excel_path="input_cal.xlsx"
book=load_workbook(excel_path)
sheet1=book.active

for row in sheet1.iter_rows(min_row=2,values_only=True):
    try:
         obj=BrfCal(row)
         obj.login()
         obj.epgupdate()
         
    except Exception as e:
        pass
     









